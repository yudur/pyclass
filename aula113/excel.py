import openpyxl
from random import uniform

pedidos = openpyxl.load_workbook("pedidos.xlsx")
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['Página1']

for linha in range(5, 16):
    planilha1.cell(linha, 1).value = linha - 1
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 200), 2)
    planilha1.cell(linha, 3).value = preco


pedidos.save('create/novo_arquivo.xlsx')

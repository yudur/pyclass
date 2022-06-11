import openpyxl
from random import uniform
from random import randrange

planilha = openpyxl.Workbook()
planilha.create_sheet("p1", 0)
planilha.create_sheet("p2", 1)

planilha1 = planilha["p1"]
planilha2 = planilha["p2"]

for linha in range(1, 16):
    planilha1.cell(linha, 1).value = linha - 1
    planilha1.cell(linha, 2).value = linha

    preco = round(uniform(10, 200), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'yudi {randrange(0, 1000, 10)}'
    planilha2.cell(linha, 2).value = linha
    planilha2.cell(linha, 2).value = round(uniform(0, 1000), 5)

planilha.save('create/file create.xlsx')
